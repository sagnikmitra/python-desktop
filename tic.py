import streamlit as st
import openpyxl as pxl
import numpy as np
import pandas as pd
import math
tic = pd.read_excel('TIC_norm_all.xlsx')
exp = pxl.load_workbook('TIC_norm_all.xlsx')
sheet = exp['Sheet1']
sheet.delete_rows(2, 1)
st.table(sheet)
# list = []
# listsimilar = []
# for i in range(3, sheet.max_row+1):
#     if((sheet.cell(row=i, column=1).value == sheet.cell(row=i, column=4).value)):
#         listsimilar.append(sheet.cell(row=i, column=1).value)
#         if(((sheet.cell(row=i, column=2).value / (sheet.cell(row=i, column=5).value)) >= 2.5) or ((sheet.cell(row=i, column=5).value / (sheet.cell(row=i, column=2).value)) >= 2.5)):
#             list.append(sheet.cell(row=i, column=1).value)
# st.write(listsimilar)
# st.write(sheet.cell(row=3, column=1).value)
# st.write(sheet.cell(row=3, column=4).value)
# max_avg_tic = float(sheet.cell(row=3, column=2).value)

# for i in range(3, sheet.max_row+1):
#     if (float(sheet.cell(row=i, column=2).value) > max_avg_tic):
#         max_avg_tic = float(sheet.cell(row=i, column=2).value)
# st.write(max_avg_tic)
